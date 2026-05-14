from __future__ import annotations

import csv
import io
import json
import logging
import unicodedata
from copy import deepcopy
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import OPENAI_MODEL, get_openai_client
from app.errors import IrisAPIError
from app.mock_data_loader import load_mock_data
from app.models.screening_event import ScreeningEvent
from app.models.user import User
from app.repositories.compliance_repository import ComplianceRepository
from app.repositories.underwriting_repository import UnderwritingRepository
from app.services.underwriting_service import UnderwritingService


logger = logging.getLogger(__name__)
UTC = timezone.utc

DEFAULT_WATCHLISTS = ["OFAC SDN", "FinCEN 314(a)"]
WORKSPACE_TRIGGER_TYPES = {"onboarding", "adhoc", "periodic", "cession_file"}
SCREENING_CACHE_COLUMN_HEADERS = [
    "entity_name",
    "aliases",
    "list_identifier",
    "entity_type",
    "country",
    "street_address",
    "city",
    "postal_code",
    "tax_identification_number",
    "company_registration_number",
    "dob",
]
COMPLIANCE_OVERRIDES_FILE = Path(__file__).resolve().parent.parent / "mock_data" / "compliance_overrides.json"
SCREENING_CACHE_LABELS = {
    "OFAC SDN": "OFAC",
    "FinCEN 314(a)": "FinCEN",
}
QUARTERLY_SCREENING_HISTORY = [
    {"period": "2025 Q2", "screened_on": "2025-04-18"},
    {"period": "2025 Q3", "screened_on": "2025-08-07"},
    {"period": "2025 Q4", "screened_on": "2025-11-14"},
]

# MOCK IMPLEMENTATION: cedent street/city/postal identifiers are not part of the current cedents schema.
# These values enrich sanctions verification without inventing new DB columns.
CEDENT_IDENTITY_MOCKS: dict[str, dict[str, str]] = {
    "CED-0991": {
        "street_address": "10 Fenchurch Avenue",
        "city": "London",
        "postal_code": "EC3M 5BN",
        "tax_identification_number": "GB-991-384-220",
        "uk_company_registration_number": "09114320",
    },
    "CED-1042": {
        "street_address": "22 Bishopsgate",
        "city": "London",
        "postal_code": "EC2N 4BQ",
        "tax_identification_number": "GB-1042-774-921",
        "uk_company_registration_number": "04281042",
    },
    "CED-1087": {
        "street_address": "Rue du Rhone 14",
        "city": "Geneva",
        "postal_code": "1204",
        "tax_identification_number": "CHE-1087.442.119",
        "uk_company_registration_number": "N/A",
    },
    "CED-1133": {
        "street_address": "200 Liberty Street",
        "city": "New York",
        "postal_code": "10281",
        "tax_identification_number": "US-TIN-13-1133110",
        "uk_company_registration_number": "N/A",
    },
    "CED-1156": {
        "street_address": "100 King Street West",
        "city": "Toronto",
        "postal_code": "M5X 1A9",
        "tax_identification_number": "CA-BN-1156-8821",
        "uk_company_registration_number": "N/A",
    },
    "CED-1201": {
        "street_address": "Leopoldstrasse 20",
        "city": "Munich",
        "postal_code": "80802",
        "tax_identification_number": "DE-1201-442-008",
        "uk_company_registration_number": "N/A",
    },
}


class ComplianceService:
    def __init__(self, repository: ComplianceRepository) -> None:
        self.repository = repository
        self.underwriting_service = UnderwritingService(UnderwritingRepository(repository.db))

    def get_sanctions_overview(self) -> dict[str, Any]:
        logger.info("Loading sanctions screening workspace")
        cases = [self._build_case_row(event) for event in self._workspace_events()]
        counts = self._workspace_counts(cases)
        return {
            "title": "Sanction Screening",
            "subtitle": "Entity resolution, risk scoring and explainable decisioning powered by IRiS",
            "workspace_note": (
                "IRiS will retrieve raw matches across OFAC, FinCEN and the available screening caches, "
                "perform entity resolution, score risk and recommend an action."
            ),
            "kpis": self._workspace_kpis(cases, counts),
            "tabs": [
                {"key": "all_cases", "label": "All Cases", "count": counts["all_cases"]},
                {"key": "pending_review", "label": "Pending Review", "count": counts["pending_review"]},
                {"key": "auto_cleared", "label": "Auto-Cleared", "count": counts["auto_cleared"]},
                {"key": "blocked", "label": "High-Risk / Blocked", "count": counts["blocked"]},
                {"key": "historical", "label": "Historical", "count": counts["historical"]},
            ],
            "filters": {
                "trigger_options": ["All", "Onboarding", "Ad-hoc", "Periodic", "Cession File"],
                "country_options": ["All", "UK", "US", "CH", "DE", "CA"],
            },
            "screening_cache_workbooks": self._serialize_screening_cache_workbooks(DEFAULT_WATCHLISTS),
            "cases": cases,
        }

    def list_active_hits(self) -> dict[str, Any]:
        logger.info("Loading active sanctions screening hits")
        items = [self._build_active_hit(event) for event in self._workspace_events() if self._ui_status(event) in {"Pending Review", "Blocked"}]
        return {"total": len(items), "items": items}

    def get_screening_case(self, screening_ref: str) -> dict[str, Any]:
        logger.info("Loading sanctions screening case detail")
        logger.debug("Sanctions case detail screening_ref=%s", screening_ref)
        event = self._get_event_or_error(screening_ref)
        context = self._case_context(screening_ref)
        case_row = self._build_case_row(event)
        entity_context = self._build_entity_section(event, context)
        matches = self._screening_matches_for_event(event, context, entity_context)
        raw_match = context.get("raw_match") or self._build_raw_match_payload(event, entity_context)
        analysis = context.get("analysis") or self._build_analysis_payload(event, raw_match)
        detail = {
            "screening_ref": screening_ref,
            "entity_name": event.entity_name,
            "title": event.entity_name,
            "subtitle": f"{self._display_trigger(event.trigger_type)} screening · {event.entity_name}",
            "status": case_row["status"],
            "trigger": self._display_trigger(event.trigger_type),
            "watchlists_screened": context.get("watchlists_screened") or case_row["watchlists"],
            "started_at": self._isoformat(event.created_at),
            "processing_seconds": context.get("processing_seconds") or 0.0,
            "entity_under_screening": entity_context,
            "summary": context.get("summary") or self._build_summary_payload(event, case_row["watchlists"]),
            "raw_match": raw_match,
            "analysis": analysis,
            "network_analysis": self._build_network_analysis(event, context, matches),
            "decision_history": self._build_decision_history(event, context, case_row["watchlists"], matches),
            "adverse_media": self._build_adverse_media(event, matches, case_row["watchlists"]),
            "audit_trail": context.get("audit_trail") or self._build_audit_trail(event, analysis),
        }
        return detail

    def screen_entity(
        self,
        screening_ref: str | None,
        entity_name: str,
        dob: str | None,
        cedent_id: str | None,
        member_id: str | None,
        trigger_type: str | None,
        cession_file_id: str | None,
        country: str | None,
        registration_number: str | None,
        aliases: str | None,
        registered_address: str | None,
        beneficial_owners: str | None,
        bank_details: str | None,
        sources: list[str] | None,
        persist_case: bool,
        actor: User,
    ) -> dict[str, Any]:
        logger.info("Running single-entity sanctions screening")
        logger.debug(
            "Sanctions screen screening_ref=%s entity_name=%s persist_case=%s trigger_type=%s cedent_id=%s sources=%s",
            screening_ref,
            entity_name,
            persist_case,
            trigger_type,
            cedent_id,
            sources,
        )
        identity_context = self._entity_identity_context(
            cedent_id=cedent_id,
            entity_name=entity_name,
            country=country,
            registration_number=registration_number,
            registered_address=registered_address,
            bank_details=bank_details,
        )
        alias_tokens = self._parse_csv_tokens(aliases)
        owners = self._parse_csv_tokens(beneficial_owners)
        normalized_sources = self._normalize_watchlist_sources(sources)
        cache_entries = self._watchlist_cache_entries(normalized_sources)
        matches = self._match_watchlists(entity_name, alias_tokens, identity_context, cache_entries)
        effective_screening_ref = screening_ref or (self._next_case_ref() if persist_case else self._legacy_screening_ref(entity_name, member_id))
        logger.debug("Screening effective_ref=%s matches=%s", effective_screening_ref, len(matches))

        if not matches:
            result_payload = {
                "screening_ref": effective_screening_ref,
                "entity_name": entity_name,
                "result": "cleared",
                "matched_lists": [],
                "llm_called": False,
                "llm_confidence": 1.0,
                "llm_reasoning": "No keyword or fuzzy watchlist match exceeded the screening threshold.",
                "method": "keyword_no_match",
                "identity_context": identity_context,
                "identity_match_summary": [],
            }
        else:
            matched_list_names = list(dict.fromkeys(item["list_name"] for item in matches))
            llm_result = self._verify_watchlist_match(entity_name, dob, matches, identity_context)
            result_payload = {
                "screening_ref": effective_screening_ref,
                "entity_name": entity_name,
                "result": "review" if llm_result["is_genuine_match"] else "cleared",
                "matched_lists": matched_list_names,
                "llm_called": llm_result["llm_called"],
                "llm_confidence": llm_result["confidence"],
                "llm_reasoning": llm_result["reasoning"],
                "method": self._screening_method_label(llm_result),
                "source": matches[0]["source"],
                "identity_context": identity_context,
                "identity_match_summary": llm_result.get("identity_match_summary", []),
            }

        if persist_case:
            self._persist_screening_case(
                screening_ref=effective_screening_ref,
                entity_name=entity_name,
                trigger_type=trigger_type or "adhoc",
                cedent_id=cedent_id,
                member_id=member_id,
                cession_file_id=cession_file_id,
                matches=matches,
                screening_result=result_payload,
                aliases=alias_tokens,
                beneficial_owners=owners,
                country=country,
                registration_number=registration_number,
                registered_address=registered_address,
                bank_details=bank_details,
                sources_screened=normalized_sources,
                actor=actor,
            )

        return {
            **result_payload,
            "trigger_type": trigger_type or "adhoc",
            "cedent_id": cedent_id,
            "member_id": member_id,
            "cession_file_id": cession_file_id,
            "sources_screened": normalized_sources,
            "persist_case": persist_case,
        }

    def get_cedent_screening(self, cedent_id: str) -> dict[str, Any]:
        logger.info("Loading sanctions screening detail for cedent")
        logger.debug("Sanctions detail cedent_id=%s", cedent_id)
        cedent_detail = self.underwriting_service.get_cedent_detail(cedent_id)
        screening = cedent_detail.get("sanction_screening")
        if screening is None:
            logger.error("Cedent screening detail missing for cedent_id=%s", cedent_id)
            raise IrisAPIError(404, "Invalid cedent ID", "Cedent screening detail not found")
        return {
            "cedent_id": cedent_id,
            "cedent_name": cedent_detail.get("legal_entity_name"),
            "screening_status": cedent_detail.get("screening_status"),
            "sanction_screening": screening,
            "identity_context": self._cedent_identity_context(cedent_id, cedent_detail.get("legal_entity_name")),
        }

    def trigger_bulk_screening(self, sources: list[str], scope: str) -> dict[str, Any]:
        logger.info("Triggering bulk sanctions screening")
        logger.debug("Bulk screening sources=%s scope=%s", sources, scope)
        normalized_sources = sources or DEFAULT_WATCHLISTS
        if scope != "all_active":
            logger.error("Bulk screening rejected because scope=%s is unsupported", scope)
            raise IrisAPIError(400, "Invalid scope", "Only all_active scope is supported in the POC")

        screened_ids: list[str] = []
        for cedent in self.repository.list_active_cedents():
            self.underwriting_service.trigger_sanction_screening(cedent.cedent_id, normalized_sources)
            screened_ids.append(cedent.cedent_id)

        return {
            "screening_run_id": f"bulk-scr-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}",
            "scope": scope,
            "sources": normalized_sources,
            "cedents_screened": len(screened_ids),
            "cedent_ids": screened_ids,
            "status": "completed",
        }

    def queue_bulk_screening(self, sources: list[str], scope: str) -> dict[str, Any]:
        logger.info("Queueing bulk sanctions screening job")
        logger.debug("Bulk screening queue sources=%s scope=%s", sources, scope)
        normalized_sources = sources or DEFAULT_WATCHLISTS
        if scope != "all_active":
            logger.error("Bulk screening queue rejected because scope=%s is unsupported", scope)
            raise IrisAPIError(400, "Invalid scope", "Only all_active scope is supported in the POC")

        job_id = f"scr-job-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}"
        self._store_last_bulk_job(
            {
                "job_id": job_id,
                "status": "queued",
                "scope": scope,
                "sources": normalized_sources,
                "estimated_duration_seconds": 120,
                "queued_at": self._isoformat(datetime.now(UTC)),
            }
        )
        return {"job_id": job_id, "status": "queued", "estimated_duration_seconds": 120}

    def resolve_hit(self, screening_ref: str, action: str, notes: str | None, reviewer: User) -> dict[str, Any]:
        logger.info("Resolving sanctions screening hit")
        logger.debug("Sanctions hit resolve screening_ref=%s action=%s reviewer=%s", screening_ref, action, reviewer.email)
        if action not in {"clear", "escalate", "mark_false_positive"}:
            logger.error("Sanctions hit resolve rejected because action=%s is invalid", action)
            raise IrisAPIError(400, "Invalid action", "Action must be clear, escalate, or mark_false_positive")

        event = self._get_event_or_error(screening_ref)
        now = datetime.now(UTC)
        result_map = {
            "clear": ("cleared", "cleared"),
            "escalate": ("escalated", "escalated"),
            "mark_false_positive": ("false_positive", "false_positive"),
        }
        event.result, event.review_outcome = result_map[action]
        event.reviewed_by = reviewer.id
        event.reviewed_at = now
        event.review_notes = notes or ""
        event.updated_at = now
        self.repository.update_screening_event(event)
        self._append_audit_entry(screening_ref, reviewer.email, "Human", self._resolution_label(action))
        return {
            "screening_ref": screening_ref,
            "status": self._ui_status(event),
            "action": action,
            "resolved_at": self._isoformat(now),
            "notes": notes,
        }

    def export_screening_report(self) -> dict[str, Any]:
        logger.info("Exporting screening report payload")
        workspace = self.get_sanctions_overview()
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Case ID", "Entity", "Trigger", "Risk", "Confidence", "Status", "Started At"])
        for item in workspace["cases"]:
            writer.writerow(
                [
                    item["screening_ref"],
                    item["entity_name"],
                    item["trigger"],
                    f'{item["risk_label"]} {item["risk_score"]}',
                    item["confidence_pct"],
                    item["status"],
                    item["started_at"],
                ]
            )
        return {
            "filename": "sanctions-screening-report.csv",
            "content_type": "text/csv;charset=utf-8;",
            "content": buffer.getvalue(),
        }

    def get_screening_cache_workbooks(self) -> dict[str, Any]:
        logger.info("Loading editable sanctions cache workbooks")
        return {"items": self._serialize_screening_cache_workbooks(DEFAULT_WATCHLISTS)}

    def update_screening_cache_workbook(self, list_name: str, entries: list[dict[str, Any]], actor: User) -> dict[str, Any]:
        logger.info("Updating sanctions cache workbook")
        logger.debug("Screening cache workbook update list_name=%s entry_count=%s actor=%s", list_name, len(entries), actor.email)
        item = self.repository.get_screening_cache_list(list_name)
        if item is None:
            logger.error("Screening cache workbook update failed because list_name=%s was not found", list_name)
            raise IrisAPIError(404, "Screening list not found", "The requested screening cache list does not exist")

        normalized_entries = [self._normalize_screening_cache_entry(entry) for entry in entries]
        now = datetime.now(UTC)
        item.data_payload = {"mock": True, "entries": normalized_entries}
        item.record_count = len(normalized_entries)
        item.last_sync = now
        item.updated_at = now
        updated = self.repository.update_screening_cache_list(item)
        return self._serialize_screening_cache_workbook(updated)

    def download_screening_cache_workbook(self, list_name: str) -> dict[str, Any]:
        logger.info("Downloading sanctions cache workbook")
        logger.debug("Screening cache workbook download list_name=%s", list_name)
        item = self.repository.get_screening_cache_list(list_name)
        if item is None:
            logger.error("Screening cache workbook download failed because list_name=%s was not found", list_name)
            raise IrisAPIError(404, "Screening list not found", "The requested screening cache list does not exist")

        workbook_bytes = self._build_screening_cache_workbook_bytes(item)
        return {
            "filename": self._screening_cache_filename(item.list_name),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "content": workbook_bytes,
        }

    def _workspace_events(self) -> list[ScreeningEvent]:
        return [event for event in self.repository.list_screening_events() if event.trigger_type in WORKSPACE_TRIGGER_TYPES]

    def _workspace_counts(self, cases: list[dict[str, Any]]) -> dict[str, int]:
        return {
            "all_cases": len(cases),
            "pending_review": sum(1 for item in cases if item["status"] == "Pending Review"),
            "auto_cleared": sum(1 for item in cases if item["status"] == "Auto-Cleared"),
            "blocked": sum(1 for item in cases if item["status"] == "Blocked"),
            "historical": sum(1 for item in cases if item["status"] == "Auto-Cleared"),
        }

    def _workspace_kpis(self, cases: list[dict[str, Any]], counts: dict[str, int]) -> list[dict[str, Any]]:
        total_cases = len(cases) or 1
        matches_found = sum(item["matches_count"] for item in cases)
        false_positives = sum(1 for event in self._workspace_events() if event.result == "false_positive")
        avg_turnaround = sum(item["processing_seconds"] for item in cases) / len(cases) if cases else 0.0
        return [
            {"id": "screenings_today", "label": "Screenings Today", "value": len(cases), "subtitle": "+12% vs avg", "tone": "positive"},
            {"id": "matches_found", "label": "Matches Found", "value": matches_found, "subtitle": "across all lists", "tone": "default"},
            {
                "id": "auto_clear_rate",
                "label": "Auto-Clear Rate",
                "value": round((counts["auto_cleared"] / total_cases) * 100),
                "subtitle": "straight-through",
                "tone": "positive",
                "suffix": "%",
            },
            {
                "id": "false_positive_rate",
                "label": "False Positive Rate",
                "value": round((false_positives / total_cases) * 100),
                "subtitle": "decision quality",
                "tone": "negative",
                "suffix": "%",
            },
            {"id": "pending_review", "label": "Pending Review", "value": counts["pending_review"], "subtitle": "analyst queue", "tone": "default"},
            {"id": "blocked", "label": "Blocked", "value": counts["blocked"], "subtitle": "confirmed hits", "tone": "negative"},
            {
                "id": "avg_turnaround",
                "label": "Avg Turnaround",
                "value": round(avg_turnaround, 1),
                "subtitle": "end-to-end",
                "tone": "negative" if avg_turnaround > 3.0 else "positive",
                "suffix": "s",
            },
        ]

    def _build_case_row(self, event: ScreeningEvent) -> dict[str, Any]:
        context = self._case_context(event.screening_ref)
        watchlists = context.get("watchlists_screened") or [self._watchlist_label(item) for item in (event.matched_lists or [])] or ["OFAC", "FinCEN"]
        confidence = int(round((event.llm_confidence or 0.0) * 100))
        return {
            "screening_ref": event.screening_ref,
            "entity_name": event.entity_name,
            "entity_subtitle": context.get("list_subtitle") or f"{self._display_trigger(event.trigger_type)} screening · {event.entity_name}",
            "trigger": self._display_trigger(event.trigger_type),
            "watchlists": watchlists,
            "matches_count": int(context.get("matches_count", 1 if event.keyword_match else 0)),
            "risk_label": context.get("risk_label") or self._risk_label(event),
            "risk_score": int(context.get("risk_score", confidence)),
            "confidence_pct": confidence,
            "ai_recommendation": context.get("ai_recommendation") or self._recommendation_label(event),
            "status": self._ui_status(event),
            "started_at": self._isoformat(event.created_at),
            "processing_seconds": float(context.get("processing_seconds", 0.0)),
            "country": context.get("country", ""),
        }

    def _build_active_hit(self, event: ScreeningEvent) -> dict[str, Any]:
        case_row = self._build_case_row(event)
        context = self._case_context(event.screening_ref)
        raw_match = context.get("raw_match") or self._build_raw_match_payload(event, self._build_entity_section(event, context))
        cedent = self.repository.get_cedent(event.cedent_id) if event.cedent_id else None
        return {
            "screening_ref": event.screening_ref,
            "hit_id": event.screening_ref,
            "entity": event.entity_name,
            "cedent_id": event.cedent_id or "",
            "cedent_name": cedent.legal_entity_name if cedent else event.entity_name,
            "member_id": event.member_id,
            "match_type": raw_match["title"].replace("Raw Match · ", "") if raw_match else (event.matched_lists or ["Watchlist"])[0],
            "source": raw_match["candidate_id"] if raw_match else ", ".join(case_row["watchlists"]),
            "confidence": round((event.llm_confidence or 0.0), 2),
            "matched_on": event.created_at.strftime("%Y-%m-%d %H:%M"),
            "status": case_row["status"],
            "reasoning": event.llm_reasoning or "",
        }

    def _persist_screening_case(
        self,
        *,
        screening_ref: str,
        entity_name: str,
        trigger_type: str,
        cedent_id: str | None,
        member_id: str | None,
        cession_file_id: str | None,
        matches: list[dict[str, Any]],
        screening_result: dict[str, Any],
        aliases: list[str],
        beneficial_owners: list[str],
        country: str | None,
        registration_number: str | None,
        registered_address: str | None,
        bank_details: str | None,
        sources_screened: list[str],
        actor: User,
    ) -> None:
        logger.info("Persisting sanctions screening case")
        logger.debug("Persist sanctions case screening_ref=%s result=%s", screening_ref, screening_result["result"])
        now = datetime.now(UTC)
        result = screening_result["result"]
        event = self.repository.get_screening_event(screening_ref)
        if event is None:
            event = ScreeningEvent(
                screening_ref=screening_ref,
                trigger_type=trigger_type,
                entity_name=entity_name,
                entity_type="organisation",
                cedent_id=cedent_id,
                member_id=member_id,
                cession_file_id=cession_file_id,
                keyword_match=bool(matches),
                matched_lists=screening_result["matched_lists"],
                llm_called=screening_result["llm_called"],
                llm_confidence=screening_result["llm_confidence"],
                llm_reasoning=screening_result["llm_reasoning"],
                llm_is_genuine=(result == "review"),
                result=result,
                review_outcome="cleared" if result == "cleared" else None,
                reviewed_at=now if result == "cleared" else None,
                reviewed_by=actor.id if result == "cleared" else None,
                review_notes="Auto-cleared by decision engine" if result == "cleared" else None,
                created_at=now,
                updated_at=now,
            )
            self.repository.create_screening_event(event)
        else:
            event.keyword_match = bool(matches)
            event.matched_lists = screening_result["matched_lists"]
            event.llm_called = screening_result["llm_called"]
            event.llm_confidence = screening_result["llm_confidence"]
            event.llm_reasoning = screening_result["llm_reasoning"]
            event.llm_is_genuine = result == "review"
            event.result = result
            event.updated_at = now
            self.repository.update_screening_event(event)

        screened_watchlists = [self._watchlist_label(item) for item in sources_screened] or ["OFAC", "FinCEN"]
        context = {
            "country": country or "",
            "registration_number": registration_number or "",
            "aliases": aliases,
            "registered_address": registered_address or "",
            "beneficial_owners": beneficial_owners,
            "bank_details": bank_details or "",
            "watchlists_screened": screened_watchlists,
            "processing_seconds": round(1.5 + (len(matches) * 0.9), 1),
            "matches_count": len(matches),
            "risk_label": self._risk_label(event),
            "risk_score": int(round((screening_result["llm_confidence"] or 0.0) * 100)) if matches else 12,
            "ai_recommendation": self._recommendation_label(event),
            "list_subtitle": f"{self._display_trigger(trigger_type)} screening · {entity_name}",
            "summary": self._build_summary_payload(event, screened_watchlists),
            "raw_match": self._build_dynamic_raw_match(matches, screening_result["identity_context"]),
            "analysis": self._build_dynamic_analysis(event, matches, screening_result),
            "audit_trail": self._build_dynamic_audit_trail(actor.email, screening_result, now),
        }
        self._store_case_context(screening_ref, context)

    def _get_event_or_error(self, screening_ref: str) -> ScreeningEvent:
        event = self.repository.get_screening_event(screening_ref)
        if event is None:
            logger.error("Sanctions screening case not found for screening_ref=%s", screening_ref)
            raise IrisAPIError(404, "Invalid screening ref", "Screening case was not found")
        return event

    def _watchlist_cache_entries(self, list_names: list[str]) -> list[dict[str, Any]]:
        entries: list[dict[str, Any]] = []
        seen_keys: set[tuple[str, str, str]] = set()

        def append_entry(list_name: str, provider: str, entry: dict[str, Any]) -> None:
            normalized_name = self._normalize_name(str(entry.get("entity_name") or ""))
            identifier = str(entry.get("list_identifier") or "").strip().lower()
            dedupe_key = (list_name, identifier, normalized_name)
            if dedupe_key in seen_keys:
                return
            seen_keys.add(dedupe_key)
            entries.append(
                {
                    **deepcopy(entry),
                    "list_name": list_name,
                    "source": provider,
                }
            )

        cache_lists = self.repository.list_screening_cache_lists(list_names)
        for item in cache_lists:
            payload = item.data_payload or {}
            for entry in payload.get("entries", []):
                append_entry(item.list_name, item.provider, entry)

        if not entries:
            logger.info("Falling back to screening cache seed payload because active cache rows do not yet contain entries")
        else:
            logger.info("Merging screening cache seed payload with active cache rows to backfill missing watchlist entries")

        for item in load_mock_data("screening_cache_lists_seed.json"):
            if item["list_name"] not in list_names:
                continue
            for entry in (item.get("data_payload") or {}).get("entries", []):
                append_entry(item["list_name"], item["provider"], entry)
        logger.debug("Loaded watchlist cache entries count=%s", len(entries))
        return entries

    def _serialize_screening_cache_workbooks(self, list_names: list[str]) -> list[dict[str, Any]]:
        workbooks: list[dict[str, Any]] = []
        cache_lists = self.repository.list_screening_cache_lists(list_names)
        fallback_seed = {item["list_name"]: item for item in load_mock_data("screening_cache_lists_seed.json")}
        for item in cache_lists:
            seed_payload = fallback_seed.get(item.list_name, {})
            workbooks.append(self._serialize_screening_cache_workbook(item, seed_payload))
        logger.debug("Serialized screening cache workbooks count=%s", len(workbooks))
        return workbooks

    def _serialize_screening_cache_workbook(self, item: Any, seed_item: dict[str, Any] | None = None) -> dict[str, Any]:
        payload = item.data_payload or {}
        seed_payload = (seed_item or {}).get("data_payload") or {}
        entries = payload.get("entries") or seed_payload.get("entries") or []
        normalized_entries = [self._normalize_screening_cache_entry(entry) for entry in entries]
        display_name = self._screening_cache_display_name(item.list_name)
        return {
            "list_name": item.list_name,
            "display_name": display_name,
            "provider": item.provider,
            "record_count": item.record_count or len(normalized_entries),
            "last_sync": self._isoformat(item.last_sync) if item.last_sync else "",
            "status": item.status,
            "filename": self._screening_cache_filename(item.list_name),
            "entries": normalized_entries,
        }

    def _normalize_screening_cache_entry(self, entry: dict[str, Any]) -> dict[str, Any]:
        aliases = entry.get("aliases", [])
        if isinstance(aliases, str):
            aliases_list = self._parse_csv_tokens(aliases)
        else:
            aliases_list = [str(item).strip() for item in aliases if str(item).strip()]
        return {
            "entity_name": str(entry.get("entity_name") or ""),
            "aliases": aliases_list,
            "list_identifier": str(entry.get("list_identifier") or ""),
            "entity_type": str(entry.get("entity_type") or ""),
            "country": str(entry.get("country") or ""),
            "street_address": str(entry.get("street_address") or ""),
            "city": str(entry.get("city") or ""),
            "postal_code": str(entry.get("postal_code") or ""),
            "tax_identification_number": entry.get("tax_identification_number") or "",
            "company_registration_number": entry.get("company_registration_number") or "",
            "dob": entry.get("dob") or "",
        }

    def _build_screening_cache_workbook_bytes(self, item: Any) -> bytes:
        payload = item.data_payload or {}
        entries = [self._normalize_screening_cache_entry(entry) for entry in payload.get("entries", [])]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._screening_cache_sheet_name(item.list_name)
        sheet["A1"] = f"{self._screening_cache_display_name(item.list_name)} sanctions cache"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A2"] = f"Provider: {item.provider}"
        sheet["B2"] = f"Records: {len(entries)}"
        sheet["C2"] = f"Last sync: {self._isoformat(item.last_sync) if item.last_sync else ''}"
        sheet["D2"] = f"Status: {item.status}"

        header_row = 4
        header_fill = PatternFill(fill_type="solid", fgColor="EEF3F7")
        for column_index, header in enumerate(SCREENING_CACHE_COLUMN_HEADERS, start=1):
            cell = sheet.cell(row=header_row, column=column_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row_index, entry in enumerate(entries, start=header_row + 1):
            row_values = [
                entry["entity_name"],
                ", ".join(entry["aliases"]),
                entry["list_identifier"],
                entry["entity_type"],
                entry["country"],
                entry["street_address"],
                entry["city"],
                entry["postal_code"],
                entry["tax_identification_number"],
                entry["company_registration_number"],
                entry["dob"],
            ]
            for column_index, value in enumerate(row_values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:K{max(len(entries) + 4, 4)}"
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 22
        sheet.column_dimensions["D"].width = 22
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 24
        sheet.column_dimensions["G"].width = 18
        sheet.column_dimensions["H"].width = 14
        sheet.column_dimensions["I"].width = 24
        sheet.column_dimensions["J"].width = 24
        sheet.column_dimensions["K"].width = 14

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _screening_cache_display_name(self, list_name: str) -> str:
        return SCREENING_CACHE_LABELS.get(list_name, list_name)

    def _screening_cache_filename(self, list_name: str) -> str:
        return f"{self._screening_cache_display_name(list_name).replace(' ', '_')}.xlsx"

    def _screening_cache_sheet_name(self, list_name: str) -> str:
        return self._screening_cache_display_name(list_name)[:31]

    def _normalize_watchlist_sources(self, sources: list[str]) -> list[str]:
        normalized: list[str] = []
        for source in sources or DEFAULT_WATCHLISTS:
            value = (source or "").strip()
            if not value:
                continue
            if value.lower() in {"ofac", "ofac sdn"}:
                value = "OFAC SDN"
            elif value.lower() in {"fincen", "fincen 314(a)", "fincen 314a"}:
                value = "FinCEN 314(a)"
            if value not in normalized:
                normalized.append(value)
        return normalized or list(DEFAULT_WATCHLISTS)

    def _match_watchlists(
        self,
        entity_name: str,
        aliases: list[str],
        identity_context: dict[str, Any] | None,
        cache_entries: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        search_terms = [entity_name, *aliases]
        matches: list[dict[str, Any]] = []
        for entry in cache_entries:
            best_score = 0.0
            for term in search_terms:
                best_score = max(best_score, self._name_similarity(term, entry["entity_name"], entry.get("aliases", [])))
            if best_score < 0.7:
                continue
            matches.append(
                {
                    **entry,
                    "score": round(best_score, 2),
                    "identity_match_summary": self._identity_match_summary(identity_context, [entry]),
                }
            )
        matches.sort(key=lambda item: item["score"], reverse=True)
        return matches[:4]

    def _entity_identity_context(
        self,
        *,
        cedent_id: str | None,
        entity_name: str,
        country: str | None,
        registration_number: str | None,
        registered_address: str | None,
        bank_details: str | None,
    ) -> dict[str, Any] | None:
        base = self._cedent_identity_context(cedent_id, entity_name) or {
            "cedent_id": cedent_id or "",
            "name": entity_name,
            "trading_name": "",
            "street_address": "",
            "city": "",
            "postal_code": "",
            "country": "",
            "ssn_tin": "",
            "uk_company_registration_number": "",
            "bank_details": "",
            "source": "request",
        }
        if registered_address:
            street, city, postal = self._split_address(registered_address)
            base["street_address"] = street or base["street_address"]
            base["city"] = city or base["city"]
            base["postal_code"] = postal or base["postal_code"]
        if country:
            base["country"] = country
        if registration_number:
            base["uk_company_registration_number"] = registration_number
        if bank_details:
            base["bank_details"] = bank_details
        return base

    def _build_entity_section(self, event: ScreeningEvent, context: dict[str, Any]) -> dict[str, Any]:
        identity_context = self._entity_identity_context(
            cedent_id=event.cedent_id,
            entity_name=event.entity_name,
            country=context.get("country"),
            registration_number=context.get("registration_number"),
            registered_address=context.get("registered_address"),
            bank_details=context.get("bank_details"),
        ) or {}
        return {
            "entity_name": event.entity_name,
            "aliases": context.get("aliases", []),
            "registration_number": context.get("registration_number", ""),
            "registered_address": context.get("registered_address", ""),
            "country": context.get("country") or identity_context.get("country", ""),
            "entity_descriptor": "Pension Fund",
            "bank_details": context.get("bank_details", ""),
            "beneficial_owners": context.get("beneficial_owners", []),
            "identity_context": identity_context,
        }

    def _build_raw_match_payload(self, event: ScreeningEvent, entity_section: dict[str, Any]) -> dict[str, Any] | None:
        matches = self._match_watchlists(
            event.entity_name,
            entity_section.get("aliases", []),
            entity_section.get("identity_context"),
            self._watchlist_cache_entries(self._event_watchlist_sources(event, {})),
        )
        return self._build_dynamic_raw_match(matches, entity_section.get("identity_context"))

    def _build_dynamic_raw_match(self, matches: list[dict[str, Any]], identity_context: dict[str, Any] | None) -> dict[str, Any] | None:
        if not matches:
            return None
        match = matches[0]
        field_scores = [
            {"label": "Name", "value": int(round(match.get("score", 0.0) * 100))},
            {
                "label": "Address",
                "value": self._field_similarity(identity_context.get("street_address", "") if identity_context else "", match.get("street_address", "")),
            },
            {
                "label": "Country",
                "value": self._field_similarity(identity_context.get("country", "") if identity_context else "", match.get("country", "")),
            },
            {"label": "Type", "value": 10 if match.get("entity_type") else 0},
        ]
        aggregate = round(sum(item["value"] for item in field_scores) / max(len(field_scores), 1))
        return {
            "title": f'Raw Match · {self._watchlist_label(match["list_name"])}',
            "candidate_name": match["entity_name"],
            "candidate_id": match.get("list_identifier") or match["list_name"],
            "subtitle": match["list_name"],
            "country": match.get("country", ""),
            "entity_type": match.get("entity_type", "Organisation"),
            "aggregate_score": aggregate,
            "field_scores": field_scores,
        }

    def _build_analysis_payload(self, event: ScreeningEvent, raw_match: dict[str, Any] | None) -> dict[str, Any]:
        confidence_pct = int(round((event.llm_confidence or 0.0) * 100))
        if event.result == "false_positive" or (event.result == "cleared" and raw_match):
            label = "Likely Different"
            headline = event.llm_reasoning or "The candidate overlap does not hold after entity resolution."
            recommended_action = "Auto-clear"
        elif event.result == "review":
            label = "Needs Review"
            headline = event.llm_reasoning or "The overlap requires analyst validation."
            recommended_action = "Review"
        else:
            label = "High Risk"
            headline = event.llm_reasoning or "The overlap should remain blocked pending manual review."
            recommended_action = "Block"
        return {
            "label": label,
            "confidence_pct": confidence_pct,
            "headline": headline,
            "body": "IRiS weighed name, jurisdiction, registration and ownership indicators against the screening caches.",
            "factors": [
                {"text": "Name alignment reviewed against OFAC / FinCEN cache", "weight_pct": min(confidence_pct, 80), "tone": "negative" if event.result in {"review", "escalated"} else "positive"},
                {"text": "Jurisdiction and registration cross-check completed", "weight_pct": 32, "tone": "positive"},
                {"text": "Ownership / banking linkage reviewed where available", "weight_pct": 18, "tone": "positive"},
            ],
            "checks": [
                {"label": "Geographic Match", "passed": event.result != "escalated"},
                {"label": "Entity Type Match", "passed": event.result != "escalated"},
                {"label": "Industry Match", "passed": event.result != "escalated"},
                {"label": "Ownership Link", "passed": event.result != "escalated"},
            ],
            "recommended_action": recommended_action,
        }

    def _build_dynamic_analysis(self, event: ScreeningEvent, matches: list[dict[str, Any]], screening_result: dict[str, Any]) -> dict[str, Any]:
        event_snapshot = ScreeningEvent(
            screening_ref=screening_result["screening_ref"],
            trigger_type="adhoc",
            entity_name=event.entity_name,
            result=screening_result["result"],
            llm_confidence=screening_result["llm_confidence"],
            llm_reasoning=screening_result["llm_reasoning"],
        )
        return self._build_analysis_payload(event_snapshot, self._build_dynamic_raw_match(matches, screening_result.get("identity_context")))

    def _build_summary_payload(self, event: ScreeningEvent, watchlists: list[str]) -> dict[str, Any]:
        watchlist_text = " · ".join(watchlists[:3]) if watchlists else "OFAC · FinCEN"
        matched_watchlists = [self._watchlist_label(str(item)) for item in (event.matched_lists or []) if str(item).strip()]
        matched_watchlist_text = " · ".join(matched_watchlists[:3]) if matched_watchlists else watchlist_text

        if not event.keyword_match:
            return {
                "headline": f"No matches across {watchlist_text}",
                "description": "Auto-cleared by Decision Engine. No analyst review required.",
                "tone": "positive",
            }
        if event.result in {"cleared", "false_positive"}:
            return {
                "headline": f"Raw match found in {matched_watchlist_text}",
                "description": "IRiS AI reviewed the raw watchlist hit and auto-cleared the entity. No analyst review required.",
                "tone": "positive",
            }
        if event.result == "review":
            return {
                "headline": f"Raw match found in {matched_watchlist_text}",
                "description": "IRiS AI retained the raw watchlist hit for compliance review. Pending analyst disposition.",
                "tone": "warning",
            }
        return {
            "headline": f"High-risk match found in {matched_watchlist_text}",
            "description": "IRiS AI retained the raw watchlist hit as high risk. Blocked pending compliance sign-off.",
            "tone": "negative",
        }

    def _build_network_analysis(self, event: ScreeningEvent, context: dict[str, Any], matches: list[dict[str, Any]]) -> dict[str, Any]:
        owners = [str(item) for item in context.get("beneficial_owners", [])]
        owner_match = any(
            self._name_similarity(owner, str(match.get("entity_name") or ""), [str(alias) for alias in match.get("aliases", [])]) >= 0.8
            for owner in owners
            for match in matches
            if str(match.get("entity_type") or "").lower() == "individual"
        )
        exact_entity_match = any(self._normalize_name(str(match.get("entity_name") or "")) == self._normalize_name(event.entity_name) for match in matches)
        corporate_match = any(str(match.get("entity_type") or "").lower() != "individual" for match in matches)

        if event.result in {"cleared", "false_positive"}:
            return {
                "items": [
                    {"label": "Parent exposure", "value": "Clear"},
                    {"label": "Subsidiary exposure", "value": "Clear"},
                    {"label": "UBO exposure", "value": "Clear"},
                ],
                "note": "Cleared against the cached OFAC / FinCEN screening lists. No linked sanctions exposure was retained in this network review.",
            }

        parent_exposure = "Blocked" if event.result == "escalated" and exact_entity_match else "Review" if exact_entity_match or corporate_match else "Clear"
        subsidiary_exposure = "Review" if corporate_match and not exact_entity_match else "Clear"
        ubo_exposure = "Blocked" if event.result == "escalated" and owner_match else "Review" if owner_match else "Clear"
        note = (
            "Network review retained linked exposure from the cached watchlist hits and requires analyst confirmation."
            if event.result == "review"
            else "Network review retained linked exposure from the cached watchlist hits and the case remains blocked pending sign-off."
        )
        return {
            "items": [
                {"label": "Parent exposure", "value": parent_exposure},
                {"label": "Subsidiary exposure", "value": subsidiary_exposure},
                {"label": "UBO exposure", "value": ubo_exposure},
            ],
            "note": note,
        }

    def _build_decision_history(
        self,
        event: ScreeningEvent,
        context: dict[str, Any],
        watchlists: list[str],
        matches: list[dict[str, Any]],
    ) -> dict[str, Any]:
        cedent = self.repository.get_cedent(event.cedent_id) if event.cedent_id else self.repository.find_cedent_by_name(event.entity_name)
        related_events = self._related_screening_events(event)
        active_contracts = [contract for contract in self.repository.list_contracts_for_cedent(cedent.cedent_id if cedent else None) if contract.status == "active"]
        watchlist_label = " / ".join(watchlists[:2]) if watchlists else "OFAC / FinCEN"

        if active_contracts:
            entries = self._build_active_contract_decision_history_entries(event, watchlist_label, bool(matches))
            note = "Active cedents show the last three quarterly sanctions refreshes for 2025 Q2, Q3 and Q4."
        elif cedent and cedent.status == "onboarding":
            screened_on = cedent.onboarded_date.isoformat() if cedent.onboarded_date else self._isoformat(event.created_at)[:10]
            entries = [
                {
                    "period": "Pre-engagement",
                    "screening_scope": "Onboarding screening",
                    "screened_on": screened_on,
                    "watchlists": watchlist_label,
                    "decision": self._decision_history_label(event.result),
                    "rationale": "Pre-engagement screening only. Quarterly re-screening begins once a contract moves to active status.",
                }
            ]
            note = "Onboarding cedents only surface the pre-engagement screen until the first active contract is in force."
        elif related_events:
            entries = [
                {
                    "period": self._period_label(item.created_at),
                    "screening_scope": self._display_trigger(item.trigger_type),
                    "screened_on": self._isoformat(item.created_at)[:10],
                    "watchlists": watchlist_label,
                    "decision": self._decision_history_label(item.result),
                    "rationale": self._decision_history_rationale(item.result, bool(matches)),
                }
                for item in reversed(related_events[-3:])
            ]
            note = "Historical sanctions decisions are replayed from prior screening activity for repeat ad-hoc entities."
        else:
            entries = [
                {
                    "period": self._period_label(event.created_at),
                    "screening_scope": self._display_trigger(event.trigger_type),
                    "screened_on": self._isoformat(event.created_at)[:10],
                    "watchlists": watchlist_label,
                    "decision": self._decision_history_label(event.result),
                    "rationale": self._decision_history_rationale(event.result, bool(matches)),
                }
            ]
            note = "No prior lifecycle history is available for this entity, so only the current screening case is shown."

        return {
            "times_reviewed": len(entries),
            "last_verdict": entries[-1]["decision"] if entries else self._decision_history_label(event.result),
            "note": note,
            "entries": entries,
        }

    def _build_active_contract_decision_history_entries(
        self,
        event: ScreeningEvent,
        watchlist_label: str,
        has_matches: bool,
    ) -> list[dict[str, str]]:
        if event.result == "escalated":
            decisions = ["Pending Review", "Pending Review", "Blocked"]
        elif event.result == "review":
            decisions = ["Cleared", "Pending Review", "Pending Review"]
        elif event.result == "false_positive":
            decisions = ["Cleared", "False Positive", "False Positive"]
        else:
            decisions = ["Cleared", "Cleared", "Cleared"] if not has_matches else ["Cleared", "False Positive", "Cleared"]

        entries: list[dict[str, str]] = []
        for index, period in enumerate(QUARTERLY_SCREENING_HISTORY):
            decision = decisions[min(index, len(decisions) - 1)]
            entries.append(
                {
                    "period": period["period"],
                    "screening_scope": "Quarterly sanctions refresh",
                    "screened_on": period["screened_on"],
                    "watchlists": watchlist_label,
                    "decision": decision,
                    "rationale": self._decision_history_rationale(self._decision_result_key(decision), has_matches),
                }
            )
        return entries

    def _build_adverse_media(
        self,
        event: ScreeningEvent,
        matches: list[dict[str, Any]],
        watchlists: list[str],
    ) -> dict[str, Any]:
        knowledge_base = load_mock_data("adverse_media_knowledge_base.json")
        entity_key = self._normalize_name(event.entity_name)
        if entity_key in knowledge_base:
            payload = deepcopy(knowledge_base[entity_key])
            payload.setdefault("sources_checked", watchlists)
            payload.setdefault("last_checked", self._isoformat(event.updated_at or event.created_at)[:10])
            payload.setdefault(
                "summary_line",
                "Recent external coverage remains available for analyst review.",
            )
            return payload

        if not matches:
            return {
                "severity": "None",
                "note": "No relevant adverse media identified in the last 24 months.",
                "summary_line": "Latest review completed with no recent negative press or enforcement coverage identified.",
                "sources_checked": watchlists,
                "last_checked": self._isoformat(event.updated_at or event.created_at)[:10],
                "records": [],
            }

        primary_sources = [self._watchlist_label(str(match.get("list_name") or "")) for match in matches[:2]]
        severity = "High" if any(source == "OFAC" for source in primary_sources) else "Medium"
        records = []
        for index, match in enumerate(matches[:2]):
            source_label = self._watchlist_label(str(match.get("list_name") or "Watchlist"))
            records.append(
                {
                    "published_at": "2025-10-07" if index == 0 else "2025-08-12",
                    "source": f"{source_label} cached watchlist",
                    "headline": f"{event.entity_name} retained a {source_label} risk-intelligence signal",
                    "summary": "No entity-specific knowledge-base profile exists, so this adverse-media entry is derived from the current sanctions and risk-screening overlap.",
                }
            )
        return {
            "severity": severity,
            "note": "Adverse-media coverage was supplemented from the screened sanctions and risk-intelligence sources because no entity-specific profile exists.",
            "summary_line": "Related external risk signals remain open and should be reviewed alongside the sanctions match.",
            "sources_checked": watchlists,
            "last_checked": self._isoformat(event.updated_at or event.created_at)[:10],
            "records": records,
        }

    def _build_audit_trail(self, event: ScreeningEvent, analysis: dict[str, Any]) -> list[dict[str, str]]:
        started_at = self._isoformat(event.created_at)
        resolved_at = self._isoformat(event.reviewed_at) if event.reviewed_at else started_at
        status_label = self._ui_status(event)
        return [
            {"timestamp": started_at, "actor": "you@reinsure.io", "actor_type": "Human", "detail": f"Initiated {self._display_trigger(event.trigger_type).lower()} screening"},
            {"timestamp": started_at, "actor": "Match Engine", "actor_type": "System", "detail": "Watchlist cache match completed"},
            {"timestamp": started_at, "actor": "IRiS", "actor_type": "AI", "detail": f"Resolved at {analysis['confidence_pct']}% confidence"},
            {"timestamp": resolved_at, "actor": "Decision Engine", "actor_type": "System", "detail": status_label},
        ]

    def _build_dynamic_audit_trail(self, actor_email: str, screening_result: dict[str, Any], now: datetime) -> list[dict[str, str]]:
        timestamp = self._isoformat(now)
        detail = "Auto-cleared" if screening_result["result"] == "cleared" else "Pending review"
        return [
            {"timestamp": timestamp, "actor": actor_email, "actor_type": "Human", "detail": "Initiated ad-hoc screening"},
            {"timestamp": timestamp, "actor": "Match Engine", "actor_type": "System", "detail": "Watchlist cache match completed"},
            {"timestamp": timestamp, "actor": "IRiS", "actor_type": "AI", "detail": f"Resolved at {int(round((screening_result['llm_confidence'] or 0.0) * 100))}% confidence"},
            {"timestamp": timestamp, "actor": "Decision Engine", "actor_type": "System", "detail": detail},
        ]

    def _event_watchlist_sources(self, event: ScreeningEvent, context: dict[str, Any]) -> list[str]:
        candidates = event.matched_lists or context.get("watchlists_screened") or DEFAULT_WATCHLISTS
        return self._normalize_watchlist_sources([str(item) for item in candidates])

    def _screening_matches_for_event(
        self,
        event: ScreeningEvent,
        context: dict[str, Any],
        entity_section: dict[str, Any],
    ) -> list[dict[str, Any]]:
        return self._match_watchlists(
            event.entity_name,
            entity_section.get("aliases", []),
            entity_section.get("identity_context"),
            self._watchlist_cache_entries(self._event_watchlist_sources(event, context)),
        )

    def _related_screening_events(self, event: ScreeningEvent) -> list[ScreeningEvent]:
        related: list[ScreeningEvent] = []
        for item in self.repository.list_screening_events():
            same_entity = self._normalize_name(item.entity_name) == self._normalize_name(event.entity_name)
            same_cedent = bool(event.cedent_id and item.cedent_id and event.cedent_id == item.cedent_id)
            if same_entity or same_cedent:
                related.append(item)
        related.sort(key=lambda item: item.created_at)
        return related

    def _decision_history_label(self, result: str) -> str:
        mapping = {
            "cleared": "Cleared",
            "false_positive": "False Positive",
            "review": "Pending Review",
            "escalated": "Blocked",
        }
        return mapping.get(result, "Pending Review")

    def _decision_result_key(self, decision: str) -> str:
        mapping = {
            "Cleared": "cleared",
            "False Positive": "false_positive",
            "Pending Review": "review",
            "Blocked": "escalated",
        }
        return mapping.get(decision, "review")

    def _decision_history_rationale(self, result: str, has_matches: bool) -> str:
        if result == "escalated":
            return "Review retained direct sanctions indicators after name, geography and ownership checks."
        if result == "review":
            return "Review found a watchlist overlap that still needs compliance analyst confirmation."
        if result == "false_positive":
            return "Review cleared the overlap after country, registration or ownership context did not align."
        if has_matches:
            return "Review retained the case as cleared after the overlap was resolved below the sanctions threshold."
        return "No OFAC or FinCEN overlap exceeded the screening threshold."

    def _period_label(self, value: datetime | None) -> str:
        if value is None:
            return "Current"
        quarter = ((value.month - 1) // 3) + 1
        return f"{value.year} Q{quarter}"

    def _read_override_store(self) -> dict[str, Any]:
        if not COMPLIANCE_OVERRIDES_FILE.exists():
            return {}
        with COMPLIANCE_OVERRIDES_FILE.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        return payload if isinstance(payload, dict) else {}

    def _write_override_store(self, payload: dict[str, Any]) -> None:
        COMPLIANCE_OVERRIDES_FILE.parent.mkdir(parents=True, exist_ok=True)
        with COMPLIANCE_OVERRIDES_FILE.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2)

    def _case_context(self, screening_ref: str) -> dict[str, Any]:
        seed_context = deepcopy(load_mock_data("screening_case_context_seed.json").get(screening_ref, {}))
        overrides = deepcopy(self._read_override_store().get("screening_cases", {}).get(screening_ref, {}))
        return {**seed_context, **overrides}

    def _store_case_context(self, screening_ref: str, payload: dict[str, Any]) -> None:
        store = self._read_override_store()
        store.setdefault("screening_cases", {})[screening_ref] = payload
        self._write_override_store(store)

    def _store_last_bulk_job(self, payload: dict[str, Any]) -> None:
        store = self._read_override_store()
        store["last_bulk_job"] = payload
        self._write_override_store(store)

    def _append_audit_entry(self, screening_ref: str, actor: str, actor_type: str, detail: str) -> None:
        context = self._case_context(screening_ref)
        audit_trail = deepcopy(context.get("audit_trail", []))
        audit_trail.append(
            {
                "timestamp": self._isoformat(datetime.now(UTC)),
                "actor": actor,
                "actor_type": actor_type,
                "detail": detail,
            }
        )
        context["audit_trail"] = audit_trail
        self._store_case_context(screening_ref, context)

    def _risk_label(self, event: ScreeningEvent) -> str:
        if event.result == "escalated":
            return "High"
        if event.result == "review":
            return "Medium"
        return "Low"

    def _recommendation_label(self, event: ScreeningEvent) -> str:
        if event.result == "escalated":
            return "Block"
        if event.result == "review":
            return "Review"
        return "Auto-clear"

    def _ui_status(self, event: ScreeningEvent) -> str:
        if event.result == "review":
            return "Pending Review"
        if event.result == "escalated":
            return "Blocked"
        return "Auto-Cleared"

    def _display_trigger(self, value: str) -> str:
        return value.replace("_", " ").replace("-", " ").title()

    def _watchlist_label(self, value: str) -> str:
        if value.startswith("OFAC"):
            return "OFAC"
        if value.startswith("FinCEN"):
            return "FinCEN"
        if value == "FIS Adverse Media":
            return "FIS Prime"
        return value

    def _next_case_ref(self) -> str:
        prefix = datetime.now(UTC).strftime("SCR-%Y-%m-")
        existing = [event.screening_ref for event in self.repository.list_screening_events() if event.screening_ref.startswith(prefix)]
        highest = 0
        for item in existing:
            try:
                highest = max(highest, int(item.split("-")[-1]))
            except ValueError:
                continue
        return f"{prefix}{highest + 1:03d}"

    def _legacy_screening_ref(self, entity_name: str, member_id: str | None) -> str:
        seed = f"{entity_name}:{member_id or ''}"
        digits = sum(ord(char) for char in seed) % 900
        return f"SHM-{digits + 100:03d}"

    def _name_similarity(self, query: str, candidate: str, aliases: list[str]) -> float:
        normalized_query = self._normalize_name(query)
        normalized_candidate = self._normalize_name(candidate)
        if not normalized_query or not normalized_candidate:
            return 0.0
        if normalized_query == normalized_candidate:
            return 1.0

        scores = [SequenceMatcher(None, normalized_query, normalized_candidate).ratio()]
        query_tokens = {token for token in self._tokenize(query) if len(token) >= 3}
        candidate_tokens = {token for token in self._tokenize(candidate) if len(token) >= 3}
        if query_tokens and candidate_tokens:
            scores.append(len(query_tokens & candidate_tokens) / max(len(query_tokens), len(candidate_tokens)))
            if query_tokens & candidate_tokens:
                scores.append(0.71)

        for alias in aliases:
            normalized_alias = self._normalize_name(alias)
            if normalized_alias and (normalized_alias == normalized_query or normalized_alias in normalized_query or normalized_query in normalized_alias):
                scores.append(0.82)
        return max(scores)

    def _field_similarity(self, left: str, right: str) -> int:
        left_normalized = self._normalize_name(left)
        right_normalized = self._normalize_name(right)
        if not left_normalized or not right_normalized:
            return 0
        return int(round(SequenceMatcher(None, left_normalized, right_normalized).ratio() * 100))

    def _tokenize(self, value: str) -> list[str]:
        ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
        return [token.lower() for token in ascii_value.replace("/", " ").replace(",", " ").split() if token.strip()]

    def _normalize_name(self, value: str) -> str:
        ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
        return "".join(char.lower() for char in ascii_value if char.isalnum())

    def _parse_csv_tokens(self, value: str | None) -> list[str]:
        if not value:
            return []
        return [item.strip() for item in value.split(",") if item.strip()]

    def _split_address(self, value: str) -> tuple[str, str, str]:
        segments = [item.strip() for item in value.split(",") if item.strip()]
        street = segments[0] if segments else ""
        city = segments[1] if len(segments) > 1 else ""
        postal = segments[2] if len(segments) > 2 else ""
        return street, city, postal

    def _isoformat(self, value: datetime | None) -> str:
        if value is None:
            return ""
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        return value.astimezone(UTC).isoformat().replace("+00:00", "Z")

    def _resolution_label(self, action: str) -> str:
        mapping = {
            "clear": "Cleared by compliance analyst",
            "escalate": "Escalated to high-risk / blocked",
            "mark_false_positive": "Marked as false positive",
        }
        return mapping[action]

    def _cedent_identity_context(self, cedent_id: str | None, entity_name: str | None = None) -> dict[str, Any] | None:
        cedent = self.repository.get_cedent(cedent_id) or self.repository.find_cedent_by_name(entity_name)
        if cedent is None:
            return None

        mock_identity = CEDENT_IDENTITY_MOCKS.get(cedent.cedent_id, {})
        tax_identifier = cedent.tax_identification_number or mock_identity.get("tax_identification_number")
        company_number = cedent.registered_company_number or mock_identity.get("uk_company_registration_number")
        return {
            "cedent_id": cedent.cedent_id,
            "name": cedent.legal_entity_name,
            "trading_name": cedent.trading_name,
            "street_address": mock_identity.get("street_address", ""),
            "city": mock_identity.get("city", ""),
            "postal_code": mock_identity.get("postal_code", ""),
            "country": cedent.country_of_registration or cedent.jurisdiction or "",
            "ssn_tin": tax_identifier or "",
            "uk_company_registration_number": company_number or "",
            "bank_details": "",
            "source": "db_plus_mock_address_overlay" if mock_identity else "db",
        }

    def _verify_watchlist_match(
        self,
        entity_name: str,
        dob: str | None,
        matches: list[dict[str, Any]],
        identity_context: dict[str, Any] | None,
    ) -> dict[str, Any]:
        conservative_exact_match = self._find_conservative_exact_match(entity_name, matches, identity_context)
        if conservative_exact_match is not None:
            logger.info(
                "Applying conservative exact watchlist screening rule entity_name=%s list_name=%s",
                entity_name,
                conservative_exact_match["list_name"],
            )
            return {
                "is_genuine_match": True,
                "confidence": 0.97,
                "reasoning": (
                    f"Exact {entity_name} entity match found in the {conservative_exact_match['list_name']} cache "
                    "with aligned jurisdiction context. Hold in pending review for compliance disposition."
                ),
                "identity_match_summary": self._identity_match_summary(identity_context, [conservative_exact_match]),
                "llm_called": False,
            }

        client = get_openai_client()
        if client is None:
            logger.info("OpenAI client is not configured; using heuristic screening verification")
            return {
                **self._fallback_llm_verification(entity_name, dob, matches, identity_context),
                "llm_called": False,
            }

        logger.info("Running OpenAI watchlist verification")
        logger.debug(
            "OpenAI screening verification entity_name=%s dob=%s matched_lists=%s",
            entity_name,
            dob,
            [item["list_name"] for item in matches],
        )
        try:
            response = client.responses.create(
                model=OPENAI_MODEL,
                instructions=(
                    "You are a sanctions-screening verification assistant for an internal compliance workflow. "
                    "Review the entity, identity context, and watchlist matches. Use name, address, country, "
                    "tax identifiers, company registration number, and aliases when available to detect genuine "
                    "matches or false positives. Return JSON only with keys is_genuine_match (boolean), "
                    "confidence (number between 0 and 1), reasoning (string), and identity_match_summary "
                    "(array of field/status objects). Use conservative compliance judgment. If there is an exact "
                    "legal-name match against a watchlist row and the jurisdiction does not conflict, prefer "
                    "pending-review treatment even when registration or address fields are sparse."
                ),
                input=json.dumps(
                    {
                        "entity_name": entity_name,
                        "date_of_birth": dob,
                        "identity_context": identity_context,
                        "matches": matches,
                    },
                    indent=2,
                ),
            )
            parsed = json.loads((response.output_text or "").strip() or "{}")
            confidence = max(0.0, min(float(parsed.get("confidence", 0.0)), 1.0))
            return {
                "is_genuine_match": bool(parsed.get("is_genuine_match")),
                "confidence": confidence,
                "reasoning": str(parsed.get("reasoning") or "OpenAI verification completed."),
                "identity_match_summary": parsed.get("identity_match_summary", []),
                "llm_called": True,
            }
        except Exception as exc:  # pragma: no cover - network/runtime path
            logger.error("OpenAI screening verification failed entity_name=%s error=%s", entity_name, exc)
            return {
                **self._fallback_llm_verification(entity_name, dob, matches, identity_context),
                "llm_called": False,
            }

    def _fallback_llm_verification(
        self,
        entity_name: str,
        dob: str | None,
        matches: list[dict[str, Any]],
        identity_context: dict[str, Any] | None,
    ) -> dict[str, Any]:
        normalized_name = self._normalize_name(entity_name)
        identity_summary = self._identity_match_summary(identity_context, matches)
        identifier_mismatch = any(
            item["status"] == "mismatch" for item in identity_summary if item["field"] in {"country", "ssn_tin", "uk_company_registration_number"}
        )
        address_mismatch = any(item["status"] == "mismatch" for item in identity_summary if item["field"] in {"city", "postal_code"})

        if normalized_name == self._normalize_name("Hans Muller"):
            if dob == "1962-03-15":
                return {
                    "is_genuine_match": True,
                    "confidence": 0.92,
                    "reasoning": "Name, DOB and list source align above the conservative review threshold. Recommend human review.",
                    "identity_match_summary": identity_summary,
                }
            return {
                "is_genuine_match": True,
                "confidence": 0.86,
                "reasoning": "Name alignment is strong, but DOB is missing or incomplete. Conservative review is still recommended.",
                "identity_match_summary": identity_summary,
            }

        if normalized_name in {self._normalize_name("Petra Schmidt"), self._normalize_name("Maple Leaf Pension Plan")}:
            return {
                "is_genuine_match": False,
                "confidence": 0.88,
                "reasoning": "Keyword match was found, but the profile lacks supporting jurisdiction or registration signals and is treated as a false positive.",
                "identity_match_summary": identity_summary,
            }

        if identifier_mismatch or address_mismatch:
            return {
                "is_genuine_match": False,
                "confidence": 0.83,
                "reasoning": (
                    "Entity-name overlap was found, but the jurisdiction, address or registration context conflicts with the watchlist record. "
                    "Treating as a likely mismatch for analyst confirmation."
                ),
                "identity_match_summary": identity_summary,
            }

        return {
            "is_genuine_match": True,
            "confidence": 0.75,
            "reasoning": f"Potential watchlist overlap found against {[item['list_name'] for item in matches]}. Human review is recommended.",
            "identity_match_summary": identity_summary,
        }

    def _find_conservative_exact_match(
        self,
        entity_name: str,
        matches: list[dict[str, Any]],
        identity_context: dict[str, Any] | None,
    ) -> dict[str, Any] | None:
        normalized_name = self._normalize_name(entity_name)
        if not normalized_name:
            return None

        identity_country = self._normalize_name(str((identity_context or {}).get("country") or ""))
        identity_tax_identifier = self._normalize_name(str((identity_context or {}).get("ssn_tin") or ""))
        identity_company_number = self._normalize_name(str((identity_context or {}).get("uk_company_registration_number") or ""))

        for item in matches:
            if self._normalize_name(str(item.get("entity_name") or "")) != normalized_name:
                continue

            match_country = self._normalize_name(str(item.get("country") or ""))
            if identity_country and match_country and identity_country != match_country:
                logger.debug(
                    "Skipping conservative exact watchlist rule because country mismatched entity_name=%s identity_country=%s match_country=%s",
                    entity_name,
                    identity_country,
                    match_country,
                )
                continue

            match_tax_identifier = self._normalize_name(str(item.get("tax_identification_number") or ""))
            if identity_tax_identifier and match_tax_identifier and identity_tax_identifier != match_tax_identifier:
                logger.debug(
                    "Skipping conservative exact watchlist rule because tax identifier mismatched entity_name=%s",
                    entity_name,
                )
                continue

            match_company_number = self._normalize_name(str(item.get("company_registration_number") or ""))
            if identity_company_number and match_company_number and identity_company_number != match_company_number:
                logger.debug(
                    "Skipping conservative exact watchlist rule because company registration mismatched entity_name=%s",
                    entity_name,
                )
                continue

            return item

        return None

    def _identity_match_summary(
        self,
        identity_context: dict[str, Any] | None,
        matches: list[dict[str, Any]],
    ) -> list[dict[str, str]]:
        if identity_context is None or not matches:
            return []

        match = matches[0]
        field_pairs = [
            ("name", identity_context.get("name"), match.get("entity_name")),
            ("street_address", identity_context.get("street_address"), match.get("street_address")),
            ("city", identity_context.get("city"), match.get("city")),
            ("postal_code", identity_context.get("postal_code"), match.get("postal_code")),
            ("country", identity_context.get("country"), match.get("country")),
            ("ssn_tin", identity_context.get("ssn_tin"), match.get("tax_identification_number")),
            ("uk_company_registration_number", identity_context.get("uk_company_registration_number"), match.get("company_registration_number")),
        ]
        summary: list[dict[str, str]] = []
        for field, left, right in field_pairs:
            left_text = str(left or "").strip()
            right_text = str(right or "").strip()
            if not left_text or not right_text or right_text.upper() == "N/A":
                status = "missing"
            elif self._normalize_name(left_text) == self._normalize_name(right_text):
                status = "match"
            else:
                status = "mismatch"
            summary.append(
                {
                    "field": field,
                    "cedent_value": left_text or "Not available",
                    "watchlist_value": right_text or "Not available",
                    "status": status,
                }
            )
        return summary

    def _screening_method_label(self, llm_result: dict[str, Any]) -> str:
        if llm_result.get("llm_called"):
            return "llm_confirmed" if llm_result["is_genuine_match"] else "llm_false_positive"
        return "heuristic_confirmed" if llm_result["is_genuine_match"] else "heuristic_false_positive"
