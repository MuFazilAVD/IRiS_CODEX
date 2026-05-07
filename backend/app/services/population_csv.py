from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


ALLOWED_POPULATION_STATUSES = {"active", "deceased", "deferred", "suspended", "transferred"}
SUPPORTED_TABULAR_UPLOAD_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


@dataclass(frozen=True)
class PopulationCsvIssue:
    row_number: int
    field_name: str
    severity: str
    issue_type: str
    description: str
    current_value: str | None = None
    ai_suggestion: str | None = None
    ai_confidence: float | None = None


@dataclass(frozen=True)
class PopulationCsvNormalizedRow:
    member_id: str
    policy_id: str | None
    date_of_birth: date
    gender: str
    smoker_status: str | None
    postcode: str | None
    annual_pension: Decimal
    pension_currency: str
    escalation_type: str | None
    escalation_rate: Decimal | None
    status: str
    date_of_death: date | None
    commencement_date: date | None
    effective_from: date | None
    verification_reference: str | None


@dataclass(frozen=True)
class PopulationCsvRowResult:
    row_number: int
    raw_data: dict[str, str]
    normalized: PopulationCsvNormalizedRow | None
    issues: list[PopulationCsvIssue]


def extract_tabular_upload_text(filename: str | None, raw_bytes: bytes) -> str:
    suffix = Path(filename or "").suffix.lower()

    if suffix in {"", ".csv"}:
        return raw_bytes.decode("utf-8-sig", errors="ignore")
    if suffix in {".xlsx", ".xlsm"}:
        return _excel_workbook_to_csv_text(raw_bytes)
    if suffix == ".xls":
        raise ValueError("Legacy Excel .xls files are not supported. Please save the workbook as .xlsx and upload again.")

    raise ValueError("Unsupported file type. Upload a CSV or Excel (.xlsx) file.")


def parse_population_file(filename: str | None, raw_bytes: bytes) -> list[PopulationCsvRowResult]:
    return parse_population_csv(extract_tabular_upload_text(filename, raw_bytes))


def parse_population_csv(content: str) -> list[PopulationCsvRowResult]:
    cleaned_content = content.lstrip("\ufeff").strip()
    if not cleaned_content:
        return []

    reader = csv.DictReader(io.StringIO(cleaned_content))
    if reader.fieldnames is None:
        raise ValueError("CSV header row is missing.")

    results: list[PopulationCsvRowResult] = []
    for row_number, row in enumerate(reader, start=1):
        raw_data = {
            _normalize_header(str(key or "")): str(value or "").strip()
            for key, value in row.items()
            if key is not None
        }
        if not any(raw_data.values()):
            continue

        issues: list[PopulationCsvIssue] = []
        member_id = _required_value(raw_data, ("member_id", "pensioner_ref"), row_number, issues)
        policy_id = _optional_value(raw_data, ("policy_id",))
        date_of_birth = _required_date(raw_data, ("date_of_birth", "dob"), row_number, issues)
        gender = _required_gender(raw_data, ("gender",), row_number, issues)
        smoker_status = _optional_value(raw_data, ("smoker_status",))
        postcode = _optional_value(raw_data, ("postcode",))
        annual_pension = _required_decimal(raw_data, ("annual_pension", "annuity_amount"), row_number, issues)
        pension_currency = (_optional_value(raw_data, ("pension_currency", "currency")) or "GBP").upper()
        escalation_type = _optional_value(raw_data, ("escalation_type", "indexation_basis"))
        escalation_rate = _optional_decimal(raw_data, ("escalation_rate",), row_number, issues)
        status = _optional_status(raw_data, ("status",), row_number, issues) or "active"
        date_of_death = _optional_date(raw_data, ("date_of_death", "death_date"), row_number, issues)
        commencement_date = _optional_date(raw_data, ("commencement_date",), row_number, issues)
        effective_from = _optional_date(raw_data, ("effective_from",), row_number, issues)
        verification_reference = _optional_value(raw_data, ("verification_reference", "verified_by"))

        has_critical = any(issue.severity == "critical" for issue in issues)
        normalized = None
        if not has_critical and member_id and date_of_birth and gender and annual_pension is not None:
            normalized = PopulationCsvNormalizedRow(
                member_id=member_id,
                policy_id=policy_id,
                date_of_birth=date_of_birth,
                gender=gender,
                smoker_status=smoker_status,
                postcode=postcode,
                annual_pension=annual_pension,
                pension_currency=pension_currency,
                escalation_type=escalation_type,
                escalation_rate=escalation_rate,
                status=status,
                date_of_death=date_of_death,
                commencement_date=commencement_date,
                effective_from=effective_from,
                verification_reference=verification_reference,
            )

        results.append(
            PopulationCsvRowResult(
                row_number=row_number,
                raw_data=raw_data,
                normalized=normalized,
                issues=issues,
            )
        )

    return results


def _excel_workbook_to_csv_text(raw_bytes: bytes) -> str:
    workbook = None
    try:
        workbook = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive against workbook corruption/parsing issues
        raise ValueError("The uploaded Excel workbook could not be read.") from exc

    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        header_row = [_spreadsheet_cell_to_text(cell) for cell in rows[0]]
        last_header_index = max((index for index, value in enumerate(header_row) if value), default=-1)
        if last_header_index < 0:
            raise ValueError("Spreadsheet header row is missing.")

        width = last_header_index + 1
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(header_row[:width])

        for row in rows[1:]:
            current_row = list(row[:width]) + [None] * max(width - len(row), 0)
            serialized_row = [_spreadsheet_cell_to_text(cell) for cell in current_row[:width]]
            if any(serialized_row):
                writer.writerow(serialized_row)

        return buffer.getvalue()
    finally:
        workbook.close()


def _spreadsheet_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def _optional_value(row: dict[str, str], aliases: tuple[str, ...]) -> str | None:
    for alias in aliases:
        value = row.get(alias)
        if value:
            return value
    return None


def _required_value(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> str | None:
    value = _optional_value(row, aliases)
    if value:
        return value
    issues.append(
        PopulationCsvIssue(
            row_number=row_number,
            field_name=aliases[0],
            severity="critical",
            issue_type="missing_required_field",
            description=f"{aliases[0]} is required.",
            current_value="",
        )
    )
    return None


def _required_date(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> date | None:
    value = _required_value(row, aliases, row_number, issues)
    if value is None:
        return None
    return _parse_date(value, aliases[0], row_number, issues)


def _optional_date(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> date | None:
    value = _optional_value(row, aliases)
    if value is None:
        return None
    return _parse_date(value, aliases[0], row_number, issues)


def _parse_date(
    value: str,
    field_name: str,
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> date | None:
    try:
        return date.fromisoformat(value)
    except ValueError:
        issues.append(
            PopulationCsvIssue(
                row_number=row_number,
                field_name=field_name,
                severity="critical",
                issue_type="invalid_date",
                description=f"{field_name} must be a valid ISO date (YYYY-MM-DD).",
                current_value=value,
            )
        )
        return None


def _required_gender(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> str | None:
    value = _required_value(row, aliases, row_number, issues)
    if value is None:
        return None
    normalized = value.strip().upper()
    if normalized in {"M", "F"}:
        return normalized
    if normalized == "MALE":
        return "M"
    if normalized == "FEMALE":
        return "F"
    issues.append(
        PopulationCsvIssue(
            row_number=row_number,
            field_name=aliases[0],
            severity="critical",
            issue_type="invalid_gender",
            description="gender must be M or F.",
            current_value=value,
        )
    )
    return None


def _required_decimal(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> Decimal | None:
    value = _required_value(row, aliases, row_number, issues)
    if value is None:
        return None
    return _parse_decimal(value, aliases[0], row_number, issues)


def _optional_decimal(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> Decimal | None:
    value = _optional_value(row, aliases)
    if value is None:
        return None
    return _parse_decimal(value, aliases[0], row_number, issues)


def _parse_decimal(
    value: str,
    field_name: str,
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> Decimal | None:
    try:
        return Decimal(value.replace(",", ""))
    except InvalidOperation:
        issues.append(
            PopulationCsvIssue(
                row_number=row_number,
                field_name=field_name,
                severity="critical",
                issue_type="invalid_number",
                description=f"{field_name} must be numeric.",
                current_value=value,
            )
        )
        return None


def _optional_status(
    row: dict[str, str],
    aliases: tuple[str, ...],
    row_number: int,
    issues: list[PopulationCsvIssue],
) -> str | None:
    value = _optional_value(row, aliases)
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized in ALLOWED_POPULATION_STATUSES:
        return normalized
    issues.append(
        PopulationCsvIssue(
            row_number=row_number,
            field_name=aliases[0],
            severity="critical",
            issue_type="invalid_status",
            description="status must be active, deceased, deferred, suspended, or transferred.",
            current_value=value,
        )
    )
    return None
